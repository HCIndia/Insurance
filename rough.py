'''file_path = "InsuranceData_Test.xlsx"

from openpyxl import load_workbook

print("i am inside this rough notebook")
# Load the Excel workbook
wb = load_workbook(file_path)

# Select the active worksheet
# Get data from row 3
sheet = wb.active

for i in range(3,16):
    row_data = [cell.value for idx, cell in enumerate(sheet[i], start=1) if idx != 6]  # sheet[3] gets all cells in row 3
    print(row_data) '''


'''from pathlib import Path

# Change 'your_directory_path' to the folder you want to search in
folder_path = Path("/Users/sukant/Documents/Python_Workspace/Python_Flask/Insurance/DataFiles")  

# Get all .mp3 files in the specified folder
xl_file = [str(file) for file in folder_path.glob("*.xlsx")]

print(xl_file)
'''

from openpyxl import load_workbook

# Load an existing Excel file
wb = load_workbook("DataFiles/Feb.xlsx")
ws = wb.active  # Select the active sheet

# Iterate through rows (values only)
for row in ws.iter_rows(min_row=3, values_only=True):
    filtered_row = row[:5] + row[6:] 
    if any(filtered_row):  # Checks if the row has at least one non-empty cell
        print(list(filtered_row))