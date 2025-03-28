from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
import logging
from typing import Final
from openpyxl import load_workbook
from pathlib import Path
from flask_cors import CORS
from datetime import datetime
import json
import pandas as pd
from sqlalchemy import or_



app = Flask(__name__)

"""
Development should be done on this branch.
Later should be merge with the main branch. 
"""

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

app.config["SQLALCHEMY_DATABASE_URI"] = 'sqlite:///personDetails.sqlite3'
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config['SQLALCHEMY_POOL_SIZE'] = 10
app.config['SQLALCHEMY_MAX_OVERFLOW'] = 5
app.config['SQLALCHEMY_POOL_RECYCLE'] = 1800
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {"pool_pre_ping": True}
app.secret_key = 'MyInsAPP'

logging.basicConfig(level=logging.DEBUG)

CORS(app)  # Enable CORS for cross-origin requests
'''app.config["SQLALCHEMY_DATABASE_URI"] = 'mysql+pymysql://Sukant:Macbook_2020@Sukant.mysql.pythonanywhere-services.com/Sukant$personDetails'
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False'''
db = SQLAlchemy(app)


class PersonDetails(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Date_of_insurance = db.Column(db.DateTime)
    Type_of_insurance = db.Column(db.String(40))
    Customer_name = db.Column(db.String(40))
    Customer_Contact_No= db.Column(db.String(40))
    Make = db.Column(db.String(40))
    Model = db.Column(db.String(40), nullable=False)
    Year_of_mfg = db.Column(db.String(120))
    Regd_No = db.Column(db.String(120))
    Regd_No_Database = db.Column(db.String(120))
    Old_ID_value = db.Column(db.String(120))
    New_ID_value = db.Column(db.String(120))
    Old_OD_value = db.Column(db.String(120))
    New_OD_value = db.Column(db.String(120))
    Old_final_premium = db.Column(db.String(120))
    New_final_premium = db.Column(db.String(120))
    Ncb = db.Column(db.String(120))
    New_NCB = db.Column(db.String(120))
    Discount = db.Column(db.String(120))
    Terms_Comp = db.Column(db.String(120))
    Terms_TP = db.Column(db.String(120))
    Insured_Company = db.Column(db.String(120))
    Insurer_Code = db.Column(db.String(120))
    New_Company = db.Column(db.String(120))
    Policy_No = db.Column(db.String(120))
    Add_Ons = db.Column(db.String(120))
    Ckyc_No = db.Column(db.String(120))
    Reference_1 = db.Column(db.String(120))
    Reference_Contact_1 = db.Column(db.String(120))
    Reference_2 = db.Column(db.String(120))
    Reference_Contact_2 = db.Column(db.String(120))
    Transfer_to = db.Column(db.String(120))

    def to_dict(self):
        """Convert SQLAlchemy object to a dictionary."""
        return {
            "id": self.id,
            "Date_of_insurance": self.Date_of_insurance.isoformat() if self.Date_of_insurance else None,  # Convert datetime to string
            "Type_of_insurance": self.Type_of_insurance,
            "Customer_name": self.Customer_name,
            "Customer_Contact_No": self.Customer_Contact_No,
            "Make": self.Make,
            "Model": self.Model,
            "Year_of_mfg": self.Year_of_mfg,
            "Regd_No": self.Regd_No,
            "Regd_No_Database": self.Regd_No_Database,
            "Old_ID_value": self.Old_ID_value,
            "New_ID_value": self.New_ID_value,
            "Old_OD_value": self.Old_OD_value,
            "New_OD_value": self.New_OD_value,
            "Old_final_premium": self.Old_final_premium,
            "New_final_premium": self.New_final_premium,
            "Ncb": self.Ncb,
            "New_NCB": self.New_NCB,
            "Discount": self.Discount,
            "Terms_Comp": self.Terms_Comp,
            "Terms_TP": self.Terms_TP,
            "Insured_Company": self.Insured_Company,
            "Insurer_Code": self.Insurer_Code,
            "New_Company": self.New_Company,
            "Policy_No": self.Policy_No,
            "Add_Ons": self.Add_Ons,
            "Ckyc_No": self.Ckyc_No,
            "Reference_1": self.Reference_1,
            "Reference_Contact_1": self.Reference_Contact_1,
            "Reference_2": self.Reference_2,
            "Reference_Contact_2": self.Reference_Contact_2,
            "Transfer_to": self.Transfer_to
        }

@app.route("/")
def start():
    return render_template("StartPage.html")

@app.route("/index",  methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        search_Person_Detail_by_reg = request.form.get("searchPersonDet")
        if (search_Person_Detail_by_reg):
            app.logger.info(f"Got the data from form action :{search_Person_Detail_by_reg}")
        #person_object = PersonDetails.query.filter_by(Regd_No=search_Person_Detail_by_reg).first()   #exact search command 
            #person_object = PersonDetails.query.filter(PersonDetails.Regd_No_Database.ilike(f'%{search_Person_Detail_by_reg}%')).all()
            person_object = PersonDetails.query.filter(
                or_(
                    PersonDetails.Regd_No_Database.ilike(f"%{search_Person_Detail_by_reg}%"),
                    PersonDetails.Customer_Contact_No.ilike(f"%{search_Person_Detail_by_reg}%"),
                    PersonDetails.Reference_Contact_1.ilike(f"%{search_Person_Detail_by_reg}%"),
                    PersonDetails.Reference_Contact_2.ilike(f"%{search_Person_Detail_by_reg}%"),
                )
            ).all()
                                                               # this is partial search
        if not person_object and search_Person_Detail_by_reg :
            #person_object = PersonDetails.query.filter(PersonDetails.Customer_name.ilike(f'%{search_Person_Detail_by_reg}%')).order_by(PersonDetails.Date_of_insurance.desc()).all() 
            person_object = PersonDetails.query.filter(
                or_(
                    PersonDetails.Customer_name.ilike(f"%{search_Person_Detail_by_reg}%"),
                    PersonDetails.Reference_1.ilike(f"%{search_Person_Detail_by_reg}%"),
                    PersonDetails.Reference_1.ilike(f"%{search_Person_Detail_by_reg}%")
                )
            ).order_by(PersonDetails.Date_of_insurance.desc()).all()
            print(f"This is from person search {person_object}")             

        if not person_object or not search_Person_Detail_by_reg :
            flash("No Records Found", "info")

        return render_template("index.html",person_object= person_object)

    return render_template("index.html")

@app.route("/fullPersonDetails", methods=['GET', 'POST'])
def fullPersonDetails():

    getFPD = request.args.get('selected')
    app.logger.info(f"FPD = {getFPD}")
    full_detail = db.session.query(PersonDetails).filter(
                            PersonDetails.id == getFPD
                        ).first()

    return render_template("FullPersonDetails.html", full_detail= full_detail)


@app.route("/searchByMonth", methods=['GET', 'POST'])
def searchByMonth():
    return render_template("SearchByMonth.html")

@app.route("/searchByMonthAndYear", methods=["GET",'POST'])
def search():
    month = request.args.get("month")
    year = request.args.get("year")
    insuranceCompany = request.args.get("insuranceCompany")
    results = None
    # if not month or not year:
    #     return jsonify({"error": "Both month and year are required"}), 400
    if month and year :
        try:
            month_number = datetime.strptime(month, "%B").month
            year = int(year)
        except ValueError:
            return jsonify({"error": "Invalid month or year"}), 400

    # Query database to filter by month and year
        results = PersonDetails.query.filter(
            db.extract('month', PersonDetails.Date_of_insurance) == month_number,
            db.extract('year', PersonDetails.Date_of_insurance) == year
        ).order_by(PersonDetails.Date_of_insurance).all()

        if insuranceCompany:
            results = [tfl for tfl in results if tfl.Insured_Company == insuranceCompany]
        else:
            session["myquery"] = [res.to_dict() for res in results]
            session['month'] = month
            session['year'] = year
        

    if not month and  not year and insuranceCompany :  # Apply filter only if insurance company is selected
        if 'myquery' in session:
            myquery = session.get('myquery')

            month = session['month']
            year= session['year'] 

            filter_result = []
            for mq in myquery:
                if mq.get("Insured_Company") == insuranceCompany:
                    filter_result.append(mq)
            return render_template("SearchByMonth.html", filter_result = filter_result,month=month,year=year )
        else:
            results = PersonDetails.query.filter(PersonDetails.Insured_Company == insuranceCompany).order_by(PersonDetails.Date_of_insurance).all()

    #print(results)
    return render_template("SearchByMonth.html", result=results, month=month,year=year )
    

@app.route("/searchByType", methods=['GET', 'POST'])
def searchByType():

    getAllTypeTuple = db.session.query(PersonDetails.Type_of_insurance).distinct().all()
    getAllType = [value[0] for value in getAllTypeTuple]
    getAllType.remove("NA")
    if request.method == "POST":
        data = request.get_json()
        type_name = data.get("month")

        result = db.session.query(PersonDetails).filter(PersonDetails.Type_of_insurance == type_name).all()
        return render_template("SearchByType.html", result = result, getAllType = getAllType)

    return render_template("SearchByType.html", getAllType = getAllType)

@app.route("/updatePerson",  methods=['POST'])
def updatePerson():

    data = request.json  # Receive edited data as a dictionary
    print("Received Data:", data)  # Debugging

    #First query the person
    person = db.session.query(PersonDetails).filter(
                            PersonDetails.id == data.get("unique_customer_id")
                        ).first()
    print(person)

    """ sample response
    {'Person Name': 'Mahabir K. Moti Hut', 'Contact No.': '765432876098765', 'Date of Insurance': '08/01/2024', 
    'Insurance Type': '4 Wheeler', 'Make': 'Maruti', 'Model': 'Ecco AC/Heater', 'Year of Manufacture': '2011', 
    'Registration No': 'BR-01-BB-2649', 'Old ID value (₹)': 'TP', 'New ID value (₹)': 'mko', 'Old OD value (₹)': '3416', 
    'New OD value (₹)': '9876', 'Old Final Premium (₹)': '4030', 'New Final Premium (₹)': 'NA98766', 'NCB (%)': 'NA',
     'Discount (%)': 'NA', 'Term COMP': 'NA', 'Term TP': 'NA', 'Insured Company': 'ROYAL SUNDARAM', 'Insurer Code': 'RB', 
     'New Company': 'NA', 'Policy No': 'UPTP462770000100', 'Add Ons': 'NA', 'CKYC No.': 'NA', 
     'Reference 1': 'C/o Carmart', 'Reference 1 Contact': 'NA', 'Reference 2': 'NA', 'Reference 2 Contact': 'NA', 'Transfer to': 'NA'}
    """

    if person:
        person.Type_of_insurance= data.get("Insurance Type")
        date_of_insurance = data.get("Date of Insurance")

        if date_of_insurance:
            date_obj = pd.to_datetime(date_of_insurance, dayfirst=True)
            person.Date_of_insurance = date_obj

        person.Customer_name = data.get("Person Name")
        person.Customer_Contact_No = data.get("Contact No.")
        person.Make = data.get("Make")
        person.Model = data.get("Model")
        person.Year_of_mfg = data.get("Year of Manufacture")
        person.Regd_No = data.get("Registration No")
        person.Regd_No_Database = "".join((data.get("Registration No")).split("-"))
        person.Old_ID_value = data.get("Old ID value (₹)")
        person.New_ID_value = data.get("New ID value (₹)")
        person.Old_OD_value = data.get("Old OD value (₹)")
        person.New_OD_value = data.get("New OD value (₹)")
        person.Old_final_premium = data.get("Old Final Premium (₹)")
        person.New_final_premium = data.get("New Final Premium (₹)")
        person.Ncb = data.get("NCB (%)")
        person.New_NCB = data.get("New NCB (%)")
        person.Discount = data.get("Discount (%)")
        person.Terms_Comp = data.get("Term COMP")
        person.Terms_TP = data.get("Term TP")
        person.Insured_Company = data.get("Insured Company")
        person.Insurer_Code = data.get("Insurer Code")
        person.New_Company = data.get("New Company")
        person.Add_Ons = data.get("Add Ons")
        person.Policy_No = data.get("Policy No","").strip()
        person.Ckyc_No = data.get("CKYC No.")
        person.Reference_1 = data.get("Reference 1")
        person.Reference_Contact_1 = data.get("Reference 1 Contact")
        person.Reference_2 = data.get("Reference 2")
        person.Reference_Contact_2 = data.get("Reference 2 Contact")
        person.Transfer_to = data.get("Transfer to")
        db.session.commit()

    db.session.refresh(person)

    #Ensure fresh data is fetched by expiring the session
    db.session.expire(person)

    person_updated = db.session.query(PersonDetails).filter(
                            PersonDetails.id == data.get("unique_customer_id")
                        ).first()
    print(person_updated.Policy_No)
    return render_template("FullPersonDetails.html", full_detail = person_updated)

@app.route("/renewPage", methods= ["POST"])
def renewPage():
    data = request.json
    key = data.get("key")
    value = data.get("value")

    print(key)
    print(value)

    record_to_renew = db.session.query(PersonDetails).filter(
                            PersonDetails.id == value).first()

    session['record_to_renew'] = record_to_renew.to_dict()

    return redirect(url_for("addPerson"))
     
@app.route("/addPerson")
def addPerson():
    file_path = "1_4 Wheeler Make, Model, Fuel Type, Variant.xlsx"
    df = pd.read_excel(file_path, sheet_name="Sheet1")
    
    df.ffill(inplace=True)  # Fill NaN values to maintain hierarchy
    df.loc[df["Type"].isin(["General Insurance", "Life Insurance", "Mediclaim"]), ["Make", "Model", "Fuel Type", "Variant"]] = "NA"
    
    dropdown_data = {}
    for _, row in df.iterrows():
        type_ = row["Type"]
        make = row["Make"]
        model = row["Model"]
        fuel_type = row["Fuel Type"]
        variant = row["Variant"]

        if type_ not in dropdown_data:
            dropdown_data[type_] = {}
        if make not in dropdown_data[type_]:
            dropdown_data[type_][make] = {}
        if model not in dropdown_data[type_][make]:
            dropdown_data[type_][make][model] = {}
        if fuel_type not in dropdown_data[type_][make][model]:
            dropdown_data[type_][make][model][fuel_type] = []

        dropdown_data[type_][make][model][fuel_type].append(variant)

    if 'record_to_renew' in session:
        person_to_renew = session['record_to_renew']
        session.clear()
        if person_to_renew:
            return render_template("AddPerson.html",
                                    dropdown_json=json.dumps(dropdown_data),
                                    person_to_renew = person_to_renew)


    return render_template("AddPerson.html",dropdown_json=json.dumps(dropdown_data))


@app.route("/deleteRecord", methods = ["POST"])
def deleteRecord():
    
    data = request.json
    key = data.get("key")
    value = data.get("value")

    print(key)
    print(value)

    record_to_delete = db.session.query(PersonDetails).filter(
                            PersonDetails.id == value).first()

    print((record_to_delete.Customer_name))
    print((record_to_delete.Customer_Contact_No))

    if record_to_delete:
        db.session.delete(record_to_delete)
        db.session.commit()
    else:
        err_message = "Unable to Delete table"
        return render_template("Delete.html", err_message = err_message)

    return render_template("Delete.html")


@app.route("/successPage", methods=['GET', 'POST'])
def SuccessPage():
    '''Here we will first chekc if the the same car number is alreday registered.
    if it already registered, display already added
    if new car then dispaly a message that you data is saved in db
    and a click bar to goto index/search page
    db operations 
    1.search , 2.save/commit'''

    """ New added Fields :
    Type_of_insurance , Customer_Contact_No , Regd_No_Database , Insurer_Code , Add_Ons , Ckyc_No , Reference_1
    Reference_Contact_1 , Reference_2 , Reference_Contact_2 ,
    """ 
    if request.method == 'POST':
        date_of_insurance = request.form.get("date_of_insurance")
        customer_name = request.form.get("customer_name")
        customer_contact_no = request.form.get("customer_contact_no")
        type_of_insurance = request.form.get("type")
        make = request.form.get("make")
        model = request.form.get("model")
        fuel = request.form.get("fuel")
        varient = request.form.get("varient")
        year_of_mfg = request.form.get("year_of_mfg")
        registration_no = request.form.get("registration_no")
        old_id_value = request.form.get("old_id_value")
        new_id_value = request.form.get("new_id_value")
        old_od_value = request.form.get("old_od_value")
        new_od_value = request.form.get("new_od_value")
        old_final_premium = request.form.get("old_final_premium")
        new_final_premium = request.form.get("new_final_premium")
        ncb = request.form.get("ncb")
        new_ncb = request.form.get("new_ncb")
        discount = request.form.get("discount")
        terms_comp = request.form.get("terms_comp")
        terms_tp = request.form.get("terms_tp")
        insured_company = request.form.get("insured_company")
        insurer_code = request.form.get("insurer_code")
        new_company = request.form.get("new_company")
        policy_no = request.form.get("policy_no","").strip()
        add_ons = request.form.getlist("add_ons")
        ckyc_no = request.form.get("ckyc_no")
        reference_1 = request.form.get("reference_1")
        reference_contact_1 = request.form.get("reference_contact_1")
        reference_2 = request.form.get("reference_2")
        reference_contact_2 = request.form.get("reference_contact_2")
        transfer_to = request.form.get("transfer_to")


        if date_of_insurance:
            date_of_insurance = datetime.strptime(date_of_insurance, "%Y-%m-%d")
        else:
            date_of_insurance = None 
        # handling Add on list
        add_ons = ",".join(add_ons)

        make = make +""+ model
        model = varient + "[" + fuel +"]"

        ''' TODO Write code to convert a string to hyphen separated string based on the reg no'''
        insert_person = PersonDetails(Date_of_insurance=date_of_insurance,Type_of_insurance =type_of_insurance,
                                    Customer_name = customer_name,Customer_Contact_No = customer_contact_no,Make=make, Model = model, Year_of_mfg = year_of_mfg,
                                    Regd_No= registration_no,Regd_No_Database= registration_no, Old_ID_value = old_id_value,New_ID_value= new_id_value,
                                    Old_OD_value=old_od_value, New_OD_value=new_od_value,
                                    Old_final_premium=old_final_premium,New_final_premium=new_final_premium,Ncb=ncb,New_NCB=new_ncb,
                                    Discount=discount,Terms_Comp=terms_comp, Terms_TP = terms_tp, Insured_Company=insured_company,Insurer_Code = insurer_code,
                                    New_Company=new_company,Policy_No=policy_no,Add_Ons = add_ons, Ckyc_No= ckyc_no, 
                                    Reference_1= reference_1, Reference_Contact_1 = reference_contact_1, Reference_2 = reference_2,
                                    Reference_Contact_2 = reference_contact_2 , Transfer_to=transfer_to)
        
        db.session.add(insert_person)
        app.logger.info("One Person is Succesfully added in Database")
        db.session.commit()
        app.logger.info("Changes have been comitted")

    return render_template("Success.html")
    


@app.route("/loadData")
def loadData():
    
    folder_path = Path("/Users/sukant/Documents/Python_Workspace/Python_Flask/Insurance/DataFiles")  
    #folder_path = Path("/home/Sukant/basic/DataFiles")

    xl_file = [str(file) for file in folder_path.glob("*.xlsx")]

    for file in xl_file:

        wb = load_workbook(file)
        ws = wb.active  # Select the active sheet

        # Iterate through rows (values only)
        for row_data in ws.iter_rows(min_row=3, values_only=True):
            if any(row_data):  # Checks if the row has at least one non-empty cell

                insert_person = PersonDetails(Date_of_insurance=row_data[0],Type_of_insurance=row_data[1],
                                            Customer_name = row_data[2],Customer_Contact_No=row_data[3],Make=row_data[4], Model = row_data[5], Year_of_mfg = row_data[6],
                                            Regd_No= row_data[7],Regd_No_Database= row_data[8], Old_ID_value = row_data[9],New_ID_value= row_data[10],
                                            Old_OD_value=row_data[11], New_OD_value=row_data[12],
                                            Old_final_premium=row_data[13],New_final_premium=row_data[14],Ncb=row_data[15],
                                            Discount=row_data[16],Terms_Comp=row_data[17],Terms_TP=row_data[18], Insured_Company=row_data[19],
                                            Insurer_Code=row_data[20] , New_Company=row_data[21],Policy_No=row_data[22], 
                                            Add_Ons= row_data[23],Ckyc_No = row_data[24] ,Reference_1 = row_data[25], Reference_Contact_1= row_data[26],
                                             Reference_2= row_data[27],
                                            Reference_Contact_2= row_data[28] , Transfer_to=row_data[29])
                #BR01BB2649
                db.session.add(insert_person)
                app.logger.info("Person is Succesfully added in Database")
                db.session.commit()
                app.logger.info("Changes have been comitted")

    return "Data is succesfully uploaded"


if __name__ == "__main__":
    with app.app_context():
        
        db.drop_all()
        db.create_all()
        app.run(debug=True, host = "0.0.0.0")


